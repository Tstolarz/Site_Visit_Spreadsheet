import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill
import gspread
from google.oauth2.service_account import Credentials
import os

# Target sites to analyze
TARGET_SITES = [
    'NANT', 'BLCK', 'AMAG', 'MRCH', 'HEMP', 'HOOK', 'LOVE',
    'BRIG', 'WILD', 'SILD', 'OLDB', 'PORT', 'CAPE', 'LEWE', 'HLPN', 
    'SEAB', 'BRAD', 'SPRK', 'HLGT', 'BRMR', 'RATH', 'WOOD', 'CMPT'
]

# Google Sheets configuration
SPREADSHEET_URL = 'https://docs.google.com/spreadsheets/d/1d5zPBR64RKGYmP6SnsuUYg4DJnLJkI7W4FdA7DniJ00/edit?usp=sharing'

def get_script_directory():
    """Get the directory where this script is located"""
    return os.path.dirname(os.path.abspath(__file__))

def find_service_account_json():
    """Find the service account JSON file in the script directory"""
    script_dir = get_script_directory()
    
    # Look for common JSON file names
    json_name = ['dazzling-howl-465916-a1-3199e1fa2c8d.json']

    # Also look for any .json file
    json_files = [f for f in os.listdir(script_dir) if f.endswith('.json')]
    
    # Check specific names first
    for name in json_name:
        json_path = os.path.join(script_dir, name)
        if os.path.exists(json_path):
            print(f"Found service account file: {name}")
            return json_path
    
    # If no specific names found, use the first JSON file found
    if json_files:
        json_path = os.path.join(script_dir, json_files[0])
        print(f"Using JSON file: {json_files[0]}")
        return json_path
    
    # No JSON file found
    print(f"No service account JSON file found in {script_dir}")
    print("Please place your service account JSON file in the same directory as this script")
    print("Expected filenames: service-account-key.json, credentials.json, or any .json file")
    return None
def load_data_from_google_sheets(spreadsheet_url, service_account_file, worksheet_name=None):
    """
    Load data directly from Google Sheets using service account
    
    Args:
        spreadsheet_url: The Google Sheets URL
        service_account_file: Path to your service account JSON file
        worksheet_name: Name of specific worksheet (optional, uses first sheet if None)
    
    Returns:
        pandas DataFrame with the data
    """
    print("Connecting to Google Sheets...")
    
    # Set up credentials and authorize
    scopes = [
        'https://www.googleapis.com/auth/spreadsheets.readonly',
        'https://www.googleapis.com/auth/drive.readonly'
    ]
    
    credentials = Credentials.from_service_account_file(
        service_account_file, 
        scopes=scopes
    )
    
    gc = gspread.authorize(credentials)
    
    # Open the spreadsheet
    try:
        # Extract spreadsheet ID from URL
        if '/d/' in spreadsheet_url:
            spreadsheet_id = spreadsheet_url.split('/d/')[1].split('/')[0]
        else:
            spreadsheet_id = spreadsheet_url
        
        spreadsheet = gc.open_by_key(spreadsheet_id)
        print(f"Successfully opened spreadsheet: {spreadsheet.title}")
        
        # Get the worksheet
        if worksheet_name:
            worksheet = spreadsheet.worksheet(worksheet_name)
        else:
            worksheet = spreadsheet.get_worksheet(0)  # First sheet
        
        print(f"Reading data from worksheet: {worksheet.title}")
        
        # Get all records as a list of dictionaries
        records = worksheet.get_all_records()
        
        # Convert to DataFrame
        df = pd.DataFrame(records)
        
        print(f"Loaded {len(df)} records with columns: {list(df.columns)}")
        
        return df
        
    except Exception as e:
        print(f"Error loading data from Google Sheets: {e}")
        print("Make sure:")
        print("1. Your service account has access to the spreadsheet")
        print("2. The spreadsheet URL is correct")
        print("3. The service account JSON file path is correct")
        return None

def analyze_site_visits_from_sheets(spreadsheet_url, service_account_file, output_file='site_visit_analysis.xlsx', worksheet_name=None):
    """
    Analyze site visit data directly from Google Sheets and create updated spreadsheet
    
    Args:
        spreadsheet_url: Google Sheets URL
        service_account_file: Path to service account JSON file
        output_file: Name for output Excel file
        worksheet_name: Specific worksheet name (optional)
    """
    
    # Load data from Google Sheets
    df = load_data_from_google_sheets(spreadsheet_url, service_account_file, worksheet_name)
    
    if df is None:
        print("Failed to load data from Google Sheets")
        return None
    
    # Convert Date column to datetime
    print("Converting date column...")
    try:
        # First try automatic detection
        df['Date'] = pd.to_datetime(df['Date'], format='mixed')
    except:
        try:
            # Try common formats that might be in your sheet
            df['Date'] = pd.to_datetime(df['Date'], format='%m/%d/%y')
        except:
            try:
                df['Date'] = pd.to_datetime(df['Date'], format='%m/%d/%Y')
            except:
                # Last resort - let pandas infer the format
                df['Date'] = pd.to_datetime(df['Date'], infer_datetime_format=True)
    
    print(f"Date conversion successful. Date range: {df['Date'].min()} to {df['Date'].max()}")
    
    # Filter for target sites only
    target_data = df[df['Site'].isin(TARGET_SITES)].copy()
    
    print(f"Found {len(target_data)} records for target sites")
    
    # Get the most recent visit for each site
    latest_visits = target_data.groupby('Site')['Date'].max().reset_index()
    latest_visits.columns = ['Site', 'Last_Visit_Date']
    
    # Calculate days since last visit
    today = datetime.now()
    latest_visits['Days_Since_Visit'] = (today - latest_visits['Last_Visit_Date']).dt.days
    
    # Add visit status
    latest_visits['Visit_Required'] = latest_visits['Days_Since_Visit'] > 180
    
    # Add priority level
    def get_priority(days):
        if days > 180:
            return 'HIGH'
        elif days > 120:
            return 'MEDIUM'
        else:
            return 'LOW'
    
    latest_visits['Priority'] = latest_visits['Days_Since_Visit'].apply(get_priority)
    
    # Check for sites with no visit records
    all_target_sites = set(TARGET_SITES)
    sites_with_data = set(latest_visits['Site'].tolist())
    sites_no_data = all_target_sites - sites_with_data
    
    if sites_no_data:
        print(f"Sites with no visit records: {sorted(sites_no_data)}")
        # Add these sites as requiring visits
        for site in sites_no_data:
            new_row = {
                'Site': site,
                'Last_Visit_Date': None,
                'Days_Since_Visit': 999,
                'Visit_Required': True,
                'Priority': 'HIGH'
            }
            latest_visits = pd.concat([latest_visits, pd.DataFrame([new_row])], ignore_index=True)
    
    # Sort by priority and days since visit
    priority_order = {'HIGH': 0, 'MEDIUM': 1, 'LOW': 2}
    latest_visits['Priority_Order'] = latest_visits['Priority'].map(priority_order)
    latest_visits = latest_visits.sort_values(['Priority_Order', 'Days_Since_Visit'], ascending=[True, False])
    latest_visits = latest_visits.drop('Priority_Order', axis=1)
    
    # Create summary statistics
    summary_stats = {
        'Total Target Sites': len(latest_visits),
        'Sites Requiring Visit (>180 days)': len(latest_visits[latest_visits['Visit_Required']]),
        'Sites OK (≤180 days)': len(latest_visits[~latest_visits['Visit_Required']]),
        'High Priority (>180 days)': len(latest_visits[latest_visits['Priority'] == 'HIGH']),
        'Medium Priority (120-180 days)': len(latest_visits[latest_visits['Priority'] == 'MEDIUM']),
        'Low Priority (<120 days)': len(latest_visits[latest_visits['Priority'] == 'LOW']),
        'Average Days Since Visit': latest_visits[latest_visits['Days_Since_Visit'] < 999]['Days_Since_Visit'].mean(),
        'Maximum Days Since Visit': latest_visits[latest_visits['Days_Since_Visit'] < 999]['Days_Since_Visit'].max(),
        'Sites with No Data': len(latest_visits[latest_visits['Days_Since_Visit'] == 999])
    }
    
    # Format the output dataframe
    output_df = latest_visits.copy()
    output_df['Last_Visit_Date'] = output_df['Last_Visit_Date'].dt.strftime('%Y-%m-%d')
    output_df['Last_Visit_Date'] = output_df['Last_Visit_Date'].fillna('No Data')
    output_df['Days_Since_Visit'] = output_df['Days_Since_Visit'].replace(999, 'No Data')
    
    # Reorder columns
    column_order = ['Site', 'Last_Visit_Date', 'Days_Since_Visit', 'Priority', 'Visit_Required']
    output_df = output_df[column_order]
    
    # Create Excel file with multiple sheets
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Main analysis sheet
        output_df.to_excel(writer, sheet_name='Site Visit Analysis', index=False)
        
        # Apply color formatting to the main sheet
        workbook = writer.book
        worksheet = writer.sheets['Site Visit Analysis']
        
        # Define fill colors
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')  # Light red
        yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')  # Light yellow
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Add some padding and set the column width
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Color rows based on priority (starting from row 2 since row 1 is headers)
        for row_idx in range(2, len(output_df) + 2):
            priority = worksheet[f'D{row_idx}'].value  # Priority column is now D
            
            if priority == 'HIGH':
                fill = red_fill
            elif priority == 'MEDIUM':
                fill = yellow_fill
            else:
                continue  # LOW priority gets no fill
            
            # Apply fill to all columns in the row
            for col_idx in range(1, len(output_df.columns) + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = fill
        
        # Summary sheet
        summary_df = pd.DataFrame(list(summary_stats.items()), columns=['Metric', 'Value'])
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Sites requiring immediate attention
        priority_sites = output_df[output_df['Priority'].isin(['HIGH', 'MEDIUM'])]
        priority_sites.to_excel(writer, sheet_name='Priority Sites', index=False)
        
        # Apply color formatting to priority sites sheet
        if 'Priority Sites' in writer.sheets:
            priority_worksheet = writer.sheets['Priority Sites']
            
            # Auto-adjust column widths for priority sites sheet
            for column in priority_worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                priority_worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Apply colors to priority sites sheet
            for row_idx in range(2, len(priority_sites) + 2):
                priority = priority_worksheet[f'D{row_idx}'].value  # Priority column is now D
                
                if priority == 'HIGH':
                    fill = red_fill
                elif priority == 'MEDIUM':
                    fill = yellow_fill
                else:
                    continue
                
                for col_idx in range(1, len(priority_sites.columns) + 1):
                    priority_worksheet.cell(row=row_idx, column=col_idx).fill = fill
        
        # Priority breakdown
        priority_breakdown = latest_visits.groupby('Priority').agg({
            'Site': lambda x: ', '.join(x),
            'Days_Since_Visit': ['count', lambda x: x[x < 999].mean() if any(x < 999) else 0]
        }).round(1)
        priority_breakdown.columns = ['Sites', 'Count', 'Avg_Days']
        priority_breakdown.to_excel(writer, sheet_name='Priority Breakdown')
        
        # Auto-adjust column widths for all remaining sheets
        for sheet_name in ['Summary', 'Priority Breakdown']:
            if sheet_name in writer.sheets:
                sheet = writer.sheets[sheet_name]
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Print summary to console
    print("\n" + "="*60)
    print("SITE VISIT ANALYSIS SUMMARY")
    print("="*60)
    
    for key, value in summary_stats.items():
        if isinstance(value, float):
            print(f"{key}: {value:.1f}")
        else:
            print(f"{key}: {value}")
    
    print(f"\n🚨 HIGH PRIORITY SITES (>180 days):")
    high = latest_visits[latest_visits['Priority'] == 'HIGH']
    for _, row in high.iterrows():
        days_text = f"{row['Days_Since_Visit']} days" if row['Days_Since_Visit'] < 999 else "No data"
        print(f"   {row['Site']}: {days_text}")
    
    print(f"\n⚠️  MEDIUM PRIORITY SITES (120-180 days):")
    medium = latest_visits[latest_visits['Priority'] == 'MEDIUM']
    for _, row in medium.iterrows():
        print(f"   {row['Site']}: {row['Days_Since_Visit']} days")
    
    print(f"\n✅ LOW PRIORITY SITES (<120 days):")
    low = latest_visits[latest_visits['Priority'] == 'LOW']
    for _, row in low.iterrows():
        print(f"   {row['Site']}: {row['Days_Since_Visit']} days")
    
    print(f"\n📊 Analysis complete! Results saved to: {output_file}")
    print("="*60)
    
    return latest_visits


# Example usage
if __name__ == "__main__":
    script_dir = get_script_directory()
    export_name = 'Most_Recent_Site_Visits.xlsx'
    output_file = os.path.join(script_dir, export_name)
    spreadsheet_url = SPREADSHEET_URL
    service_account_file =   find_service_account_json()
    
    # Run the analysis
    result = analyze_site_visits_from_sheets(
        spreadsheet_url=spreadsheet_url,
        service_account_file=service_account_file,
        output_file=output_file
    )